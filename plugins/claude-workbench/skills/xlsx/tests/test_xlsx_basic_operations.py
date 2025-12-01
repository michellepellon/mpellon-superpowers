import os
import sys
import tempfile
import unittest
from pathlib import Path

import pandas as pd

sys.path.insert(0, str(Path(__file__).parent.parent / "scripts"))

from xlsx_operations import (
    create_workbook,
    read_excel,
    write_dataframe,
    add_formula,
    set_cell_color,
    ExcelOperationError,
)


class TestXLSXBasicOperations(unittest.TestCase):
    """Test basic Excel operations."""

    def setUp(self):
        """Set up test fixtures before each test method."""
        self.test_dir = tempfile.mkdtemp()

    def tearDown(self):
        """Clean up test fixtures after each test method."""
        import shutil

        if os.path.exists(self.test_dir):
            shutil.rmtree(self.test_dir)

    def test_create_workbook_success(self):
        """Test creating a new workbook."""
        output_path = os.path.join(self.test_dir, "test.xlsx")
        wb = create_workbook(output_path)

        self.assertIsNotNone(wb)
        self.assertTrue(os.path.exists(output_path))

    def test_read_excel_success(self):
        """Test reading an Excel file."""
        output_path = os.path.join(self.test_dir, "test.xlsx")

        df = pd.DataFrame({"A": [1, 2, 3], "B": [4, 5, 6]})
        df.to_excel(output_path, index=False)

        result_df = read_excel(output_path)

        self.assertIsNotNone(result_df)
        self.assertEqual(len(result_df), 3)
        self.assertEqual(list(result_df.columns), ["A", "B"])

    def test_read_excel_file_not_found(self):
        """Test reading a non-existent Excel file."""
        with self.assertRaises(ExcelOperationError) as context:
            read_excel("nonexistent.xlsx")
        self.assertIn("not found", str(context.exception).lower())

    def test_write_dataframe_success(self):
        """Test writing a DataFrame to Excel."""
        output_path = os.path.join(self.test_dir, "output.xlsx")

        df = pd.DataFrame(
            {
                "Name": ["Alice", "Bob"],
                "Age": [25, 30],
                "Salary": [50000, 60000],
            }
        )

        write_dataframe(df, output_path)

        self.assertTrue(os.path.exists(output_path))

        result_df = read_excel(output_path)
        self.assertEqual(len(result_df), 2)
        self.assertEqual(list(result_df.columns), ["Name", "Age", "Salary"])

    def test_add_formula_success(self):
        """Test adding a formula to a workbook."""
        output_path = os.path.join(self.test_dir, "formula.xlsx")

        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws["A1"] = 10
        ws["A2"] = 20

        add_formula(ws, "A3", "=SUM(A1:A2)")

        self.assertEqual(ws["A3"].value, "=SUM(A1:A2)")

        wb.save(output_path)
        wb.close()

    def test_add_formula_invalid_cell(self):
        """Test adding formula to invalid cell reference."""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active

        with self.assertRaises(ExcelOperationError) as context:
            add_formula(ws, "INVALID", "=SUM(A1:A2)")
        self.assertIn("invalid", str(context.exception).lower())

    def test_set_cell_color_blue(self):
        """Test setting cell color to blue (for inputs)."""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws["A1"] = 100

        set_cell_color(ws, "A1", "blue")

        self.assertEqual(ws["A1"].font.color.rgb, "000000FF")

    def test_set_cell_color_black(self):
        """Test setting cell color to black (for formulas)."""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws["A1"] = "=SUM(B1:B10)"

        set_cell_color(ws, "A1", "black")

        self.assertEqual(ws["A1"].font.color.rgb, "00000000")

    def test_set_cell_color_green(self):
        """Test setting cell color to green (for links)."""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws["A1"] = "=Sheet2!A1"

        set_cell_color(ws, "A1", "green")

        self.assertEqual(ws["A1"].font.color.rgb, "00008000")

    def test_set_cell_color_invalid_color(self):
        """Test setting invalid color raises error."""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active

        with self.assertRaises(ExcelOperationError) as context:
            set_cell_color(ws, "A1", "purple")
        self.assertIn("invalid", str(context.exception).lower())


class TestXLSXPandasIntegration(unittest.TestCase):
    """Test pandas integration for data analysis."""

    def setUp(self):
        """Set up test fixtures."""
        self.test_dir = tempfile.mkdtemp()

    def tearDown(self):
        """Clean up test fixtures."""
        import shutil

        if os.path.exists(self.test_dir):
            shutil.rmtree(self.test_dir)

    def test_read_multiple_sheets(self):
        """Test reading multiple sheets from Excel."""
        output_path = os.path.join(self.test_dir, "multi_sheet.xlsx")

        df1 = pd.DataFrame({"A": [1, 2]})
        df2 = pd.DataFrame({"B": [3, 4]})

        with pd.ExcelWriter(output_path) as writer:
            df1.to_excel(writer, sheet_name="Sheet1", index=False)
            df2.to_excel(writer, sheet_name="Sheet2", index=False)

        sheets = pd.read_excel(
            output_path, sheet_name=None
        )

        self.assertEqual(len(sheets), 2)
        self.assertIn("Sheet1", sheets)
        self.assertIn("Sheet2", sheets)


if __name__ == "__main__":
    unittest.main()
