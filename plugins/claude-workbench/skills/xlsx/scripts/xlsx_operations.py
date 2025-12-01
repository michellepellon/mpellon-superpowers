"""
Core Excel operations for creating, reading, and manipulating spreadsheets.

This module provides functions for common Excel operations including:
- Creating and reading workbooks
- Adding formulas
- Setting cell colors per financial model standards
- Reading/writing with pandas
- Error handling
"""

import os
from typing import Any

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


class ExcelOperationError(Exception):
    """Raised when an Excel operation fails."""

    pass


def create_workbook(file_path: str) -> Workbook:
    """
    Create a new Excel workbook and save it.

    Args:
        file_path: Path where the workbook will be saved

    Returns:
        Workbook object

    Raises:
        ExcelOperationError: If workbook cannot be created
    """
    try:
        wb = Workbook()
        wb.save(file_path)
        return wb
    except Exception as e:
        raise ExcelOperationError(
            f"Failed to create workbook: {str(e)}"
        )


def read_excel(
    file_path: str, sheet_name: str | int = 0
) -> pd.DataFrame:
    """
    Read an Excel file into a pandas DataFrame.

    Args:
        file_path: Path to the Excel file
        sheet_name: Sheet name or index to read (default: first sheet)

    Returns:
        DataFrame containing the data

    Raises:
        ExcelOperationError: If file cannot be read
    """
    if not os.path.exists(file_path):
        raise ExcelOperationError(
            f"Excel file not found: {file_path}"
        )

    try:
        df = pd.read_excel(file_path, sheet_name=sheet_name)
        return df
    except Exception as e:
        raise ExcelOperationError(
            f"Failed to read Excel file: {str(e)}"
        )


def write_dataframe(
    df: pd.DataFrame,
    file_path: str,
    sheet_name: str = "Sheet1",
    index: bool = False,
) -> None:
    """
    Write a DataFrame to an Excel file.

    Args:
        df: DataFrame to write
        file_path: Path where the file will be saved
        sheet_name: Name of the sheet (default: "Sheet1")
        index: Whether to write DataFrame index (default: False)

    Raises:
        ExcelOperationError: If write fails
    """
    try:
        df.to_excel(file_path, sheet_name=sheet_name, index=index)
    except Exception as e:
        raise ExcelOperationError(
            f"Failed to write DataFrame: {str(e)}"
        )


def add_formula(ws: Any, cell: str, formula: str) -> None:
    """
    Add a formula to a cell in a worksheet.

    Args:
        ws: Worksheet object from openpyxl
        cell: Cell reference (e.g., "A1")
        formula: Excel formula (must start with "=")

    Raises:
        ExcelOperationError: If cell reference is invalid
    """
    try:
        ws[cell] = formula
    except Exception as e:
        raise ExcelOperationError(
            f"Failed to add formula to {cell}: invalid cell reference"
        )


# Financial model color standards (RGB values from SKILL.md)
FINANCIAL_COLORS = {
    "blue": "000000FF",  # Hardcoded inputs
    "black": "00000000",  # Formulas
    "green": "00008000",  # Links within workbook
    "red": "00FF0000",  # External links
    "yellow_bg": "00FFFF00",  # Key assumptions (background)
}


def set_cell_color(ws: Any, cell: str, color: str) -> None:
    """
    Set cell font color per financial model standards.

    Args:
        ws: Worksheet object from openpyxl
        cell: Cell reference (e.g., "A1")
        color: Color name ("blue", "black", "green", "red")

    Raises:
        ExcelOperationError: If color is invalid
    """
    if color not in FINANCIAL_COLORS:
        raise ExcelOperationError(
            f"Invalid color '{color}'. Must be one of: "
            f"{', '.join(FINANCIAL_COLORS.keys())}"
        )

    try:
        rgb = FINANCIAL_COLORS[color]
        ws[cell].font = Font(color=rgb)
    except Exception as e:
        raise ExcelOperationError(
            f"Failed to set color for {cell}: {str(e)}"
        )
